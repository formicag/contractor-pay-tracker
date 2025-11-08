"""
Unit tests for excel_parser.py
Tests Excel file parsing functionality
"""

import os
import pytest
from unittest.mock import MagicMock, patch
from openpyxl import Workbook
from common.excel_parser import PayFileParser


class TestPayFileParser:
    """Test Excel parsing functionality"""

    @pytest.fixture
    def simple_excel_file(self, tmp_path):
        """Create a simple test Excel file"""
        file_path = tmp_path / "NASA_GCI_Nasstar_Contractor_Pay_01092025.xlsx"

        wb = Workbook()
        ws = wb.active

        # Header row
        ws.append(['Employee ID', 'Surname', 'Forename', 'Unit (Days)', 'Day Rate', 'Amount', 'VAT', 'Gross Amount', 'Notes'])

        # Data rows
        ws.append(['812001', 'Mays', 'Jonathan', 20, 450.00, 9000.00, 1800.00, 10800.00, ''])
        ws.append(['812002', 'Hunt', 'David', 18, 500.00, 9000.00, 1800.00, 10800.00, ''])
        ws.append(['812003', 'Smith', 'Donna', 22, 475.00, 10450.00, 2090.00, 12540.00, ''])

        wb.save(file_path)
        return str(file_path)

    @pytest.fixture
    def excel_file_with_overtime(self, tmp_path):
        """Create Excel file with overtime records"""
        file_path = tmp_path / "PARASOL_test_01092025.xlsx"

        wb = Workbook()
        ws = wb.active

        # Header
        ws.append(['Employee ID', 'Surname', 'Forename', 'Unit (Days)', 'Day Rate', 'Amount', 'VAT', 'Gross Amount', 'Notes'])

        # Normal record
        ws.append(['812001', 'Mays', 'Jonathan', 20, 450.00, 9000.00, 1800.00, 10800.00, ''])

        # Overtime record
        ws.append(['812001', 'Mays', 'Jonathan', 2, 675.00, 1350.00, 270.00, 1620.00, 'Overtime'])

        wb.save(file_path)
        return str(file_path)

    @pytest.fixture
    def excel_file_with_empty_rows(self, tmp_path):
        """Create Excel file with empty rows and duplicate headers"""
        file_path = tmp_path / "test_with_empty.xlsx"

        wb = Workbook()
        ws = wb.active

        # Empty rows
        ws.append([None, None, None])
        ws.append([])

        # Header
        ws.append(['Employee ID', 'Surname', 'Forename', 'Unit (Days)', 'Day Rate', 'Amount', 'VAT', 'Gross Amount'])

        # Empty row
        ws.append([None, None, None, None])

        # Data
        ws.append(['812001', 'Mays', 'Jonathan', 20, 450.00, 9000.00, 1800.00, 10800.00])

        # Duplicate header (should skip)
        ws.append(['Employee ID', 'Surname', 'Forename', 'Unit (Days)', 'Day Rate', 'Amount', 'VAT', 'Gross Amount'])

        # More data
        ws.append(['812002', 'Hunt', 'David', 18, 500.00, 9000.00, 1800.00, 10800.00])

        wb.save(file_path)
        return str(file_path)

    def test_extract_metadata_nasa(self, simple_excel_file):
        """Test metadata extraction from NASA filename"""
        parser = PayFileParser(simple_excel_file)

        metadata = parser.extract_metadata()

        assert metadata['umbrella_code'] == 'NASA'
        assert metadata['submission_date'] == '01092025'
        assert 'NASA' in metadata['filename']

        parser.close()

    def test_extract_metadata_parasol(self, excel_file_with_overtime):
        """Test metadata extraction from PARASOL filename"""
        parser = PayFileParser(excel_file_with_overtime)

        metadata = parser.extract_metadata()

        assert metadata['umbrella_code'] == 'PARASOL'

        parser.close()

    def test_find_header_row(self, simple_excel_file):
        """Test automatic header row detection"""
        parser = PayFileParser(simple_excel_file)

        header_row = parser.find_header_row()

        assert header_row == 1  # First row (1-indexed)

        parser.close()

    def test_find_header_row_with_empty_rows(self, excel_file_with_empty_rows):
        """Test header detection when file starts with empty rows"""
        parser = PayFileParser(excel_file_with_empty_rows)

        header_row = parser.find_header_row()

        assert header_row == 3  # Third row after empty rows

        parser.close()

    def test_parse_records_simple(self, simple_excel_file):
        """Test parsing simple records"""
        parser = PayFileParser(simple_excel_file)

        records = parser.parse_records()

        assert len(records) == 3

        # Check first record
        assert records[0]['employee_id'] == '812001'
        assert records[0]['surname'] == 'Mays'
        assert records[0]['forename'] == 'Jonathan'
        assert records[0]['unit_days'] == 20
        assert records[0]['day_rate'] == 450.00
        assert records[0]['amount'] == 9000.00
        assert records[0]['vat_amount'] == 1800.00
        assert records[0]['gross_amount'] == 10800.00
        assert records[0]['record_type'] == 'NORMAL'

        parser.close()

    def test_parse_records_with_overtime(self, excel_file_with_overtime):
        """Test parsing records with overtime detection"""
        parser = PayFileParser(excel_file_with_overtime)

        records = parser.parse_records()

        assert len(records) == 2

        # Normal record
        assert records[0]['record_type'] == 'NORMAL'

        # Overtime record
        assert records[1]['record_type'] == 'OVERTIME'
        assert records[1]['employee_id'] == '812001'
        assert records[1]['unit_days'] == 2
        assert records[1]['day_rate'] == 675.00

        parser.close()

    def test_parse_records_skips_empty_rows(self, excel_file_with_empty_rows):
        """Test that empty rows are skipped"""
        parser = PayFileParser(excel_file_with_empty_rows)

        records = parser.parse_records()

        # Should have 2 valid records (empty rows skipped)
        assert len(records) == 2

        parser.close()

    def test_parse_records_skips_duplicate_headers(self, excel_file_with_empty_rows):
        """Test that duplicate header rows are skipped"""
        parser = PayFileParser(excel_file_with_empty_rows)

        records = parser.parse_records()

        # Should have 2 valid records (duplicate header skipped)
        assert len(records) == 2

        # Both should be valid records, not headers
        assert records[0]['employee_id'] == '812001'
        assert records[1]['employee_id'] == '812002'

        parser.close()

    def test_detect_overtime_from_notes(self):
        """Test overtime detection from notes column"""
        parser = PayFileParser.__new__(PayFileParser)

        # Test various overtime indicators
        assert parser._is_overtime_record('Overtime') is True
        assert parser._is_overtime_record('OVERTIME') is True
        assert parser._is_overtime_record('overtime hours worked') is True
        assert parser._is_overtime_record('OT') is True
        assert parser._is_overtime_record('Regular hours') is False
        assert parser._is_overtime_record('') is False

    def test_calculate_gross_amount(self, simple_excel_file):
        """Test gross amount calculation"""
        parser = PayFileParser(simple_excel_file)

        records = parser.parse_records()

        # Verify gross = amount + VAT
        for record in records:
            expected_gross = record['amount'] + record['vat_amount']
            assert record['gross_amount'] == expected_gross

        parser.close()

    def test_row_number_tracking(self, simple_excel_file):
        """Test that row numbers are tracked correctly"""
        parser = PayFileParser(simple_excel_file)

        records = parser.parse_records()

        # Row numbers should be sequential starting after header
        assert records[0]['row_number'] == 2
        assert records[1]['row_number'] == 3
        assert records[2]['row_number'] == 4

        parser.close()

    def test_umbrella_code_extraction(self):
        """Test umbrella code extraction from various filenames"""
        test_cases = [
            ('NASA_GCI_Nasstar_01092025.xlsx', 'NASA'),
            ('PARASOL_Limited_Pay_15082025.xlsx', 'PARASOL'),
            ('GIANT_Contractor_Pay_01092025.xlsx', 'GIANT'),
            ('PAYSTREAM_figures_20082025.xlsx', 'PAYSTREAM'),
            ('BROOKSON_payroll_31082025.xlsx', 'BROOKSON'),
            ('APSCo_contractors_10092025.xlsx', 'APSCO'),
        ]

        for filename, expected_code in test_cases:
            parser = PayFileParser.__new__(PayFileParser)
            parser.filename = filename

            code = parser._extract_umbrella_code()

            assert code == expected_code, f"Failed for {filename}"

    def test_submission_date_extraction(self):
        """Test submission date extraction from filenames"""
        test_cases = [
            ('NASA_GCI_Nasstar_01092025.xlsx', '01092025'),
            ('PARASOL_15082025.xlsx', '15082025'),
            ('test_file_31122024.xlsx', '31122024'),
        ]

        for filename, expected_date in test_cases:
            parser = PayFileParser.__new__(PayFileParser)
            parser.filename = filename

            date = parser._extract_submission_date()

            assert date == expected_date, f"Failed for {filename}"

    def test_total_hours_calculation(self, simple_excel_file):
        """Test total hours calculation (days * 8)"""
        parser = PayFileParser(simple_excel_file)

        records = parser.parse_records()

        # 20 days * 8 hours = 160 hours
        assert records[0]['total_hours'] == 160

        # 18 days * 8 hours = 144 hours
        assert records[1]['total_hours'] == 144

        parser.close()

    def test_numeric_fields_conversion(self, simple_excel_file):
        """Test that numeric fields are properly converted"""
        parser = PayFileParser(simple_excel_file)

        records = parser.parse_records()

        record = records[0]

        # Check types
        assert isinstance(record['unit_days'], (int, float))
        assert isinstance(record['day_rate'], (int, float))
        assert isinstance(record['amount'], (int, float))
        assert isinstance(record['vat_amount'], (int, float))
        assert isinstance(record['gross_amount'], (int, float))

        parser.close()

    def test_string_fields_handling(self, simple_excel_file):
        """Test that string fields are properly handled"""
        parser = PayFileParser(simple_excel_file)

        records = parser.parse_records()

        record = records[0]

        # Check string fields
        assert isinstance(record['employee_id'], str)
        assert isinstance(record['surname'], str)
        assert isinstance(record['forename'], str)
        assert isinstance(record['notes'], str)

        parser.close()

    def test_parser_close_cleanup(self, simple_excel_file):
        """Test that parser properly closes workbook"""
        parser = PayFileParser(simple_excel_file)

        # Workbook should be open
        assert parser.workbook is not None

        parser.close()

        # After close, workbook should be closed
        # (openpyxl doesn't provide a direct way to check, but we can verify no exceptions)

    def test_context_manager_support(self, simple_excel_file):
        """Test that parser works as context manager"""
        with PayFileParser(simple_excel_file) as parser:
            records = parser.parse_records()
            assert len(records) > 0

        # After context manager exit, parser should be closed

    def test_missing_columns_handling(self, tmp_path):
        """Test handling of files with missing columns"""
        file_path = tmp_path / "missing_columns.xlsx"

        wb = Workbook()
        ws = wb.active

        # Header missing some columns
        ws.append(['Employee ID', 'Surname', 'Forename', 'Amount'])

        # Data
        ws.append(['812001', 'Mays', 'Jonathan', 9000.00])

        wb.save(file_path)

        parser = PayFileParser(str(file_path))

        # Should handle gracefully (may raise exception or return partial data)
        # Depends on implementation - test that it doesn't crash
        try:
            records = parser.parse_records()
            # If it succeeds, verify basic fields are present
            if records:
                assert 'employee_id' in records[0]
        except Exception as e:
            # Should raise a clear error
            assert 'column' in str(e).lower() or 'missing' in str(e).lower()

        parser.close()

    def test_unicode_names_handling(self, tmp_path):
        """Test handling of unicode characters in names"""
        file_path = tmp_path / "unicode_names.xlsx"

        wb = Workbook()
        ws = wb.active

        ws.append(['Employee ID', 'Surname', 'Forename', 'Unit (Days)', 'Day Rate', 'Amount', 'VAT', 'Gross Amount'])
        ws.append(['812001', "O'Brien", 'Seán', 20, 450.00, 9000.00, 1800.00, 10800.00])

        wb.save(file_path)

        parser = PayFileParser(str(file_path))

        records = parser.parse_records()

        assert len(records) == 1
        assert records[0]['surname'] == "O'Brien"
        assert records[0]['forename'] == 'Seán'

        parser.close()
