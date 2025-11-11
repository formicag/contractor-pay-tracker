"""
Canonical Groups Matching for File Uploads

Extracts company names from filenames and Excel content,
then matches against canonical customer groups.
"""

import re
import openpyxl
from io import BytesIO
from typing import List, Dict, Optional, Set
from canonical_groups_manager import canonical_groups_manager


def extract_company_from_filename(filename: str) -> List[str]:
    """
    Extract potential company names from filename.

    Common patterns:
    - "Nasstar_payfile_2024.xlsx" -> ["Nasstar"]
    - "Tesco Mobile - Invoice.xlsx" -> ["Tesco Mobile"]
    - "GCI Network Solutions Ltd payroll.xlsx" -> ["GCI Network Solutions Ltd"]

    Args:
        filename: Original filename

    Returns:
        List of potential company names extracted from filename
    """
    candidates = []

    # Remove file extension
    name_without_ext = filename.rsplit('.', 1)[0]

    # Remove common suffixes and noise words
    noise_words = [
        'payfile', 'payroll', 'invoice', 'statement', 'report',
        'monthly', 'weekly', 'payment', 'pay', 'file', 'doc', 'document',
        '2024', '2023', '2025', 'jan', 'feb', 'mar', 'apr', 'may', 'jun',
        'jul', 'aug', 'sep', 'oct', 'nov', 'dec', 'january', 'february',
        'march', 'april', 'june', 'july', 'august', 'september',
        'october', 'november', 'december'
    ]

    # Strategy 1: Split by common delimiters (underscore, dash, space)
    for delimiter in ['_', '-', ' ']:
        parts = name_without_ext.split(delimiter)
        for part in parts:
            part_clean = part.strip()

            # Skip if empty, too short, or is a noise word
            if not part_clean or len(part_clean) < 2:
                continue
            if part_clean.lower() in noise_words:
                continue
            if part_clean.isdigit():
                continue

            candidates.append(part_clean)

    # Strategy 2: Try full filename without extension (cleaned)
    full_clean = re.sub(r'[_-]', ' ', name_without_ext)
    full_clean = re.sub(r'\s+', ' ', full_clean).strip()
    if full_clean and len(full_clean) > 2:
        candidates.append(full_clean)

    # Strategy 3: Look for multi-word company names (2-4 words)
    words = name_without_ext.replace('_', ' ').replace('-', ' ').split()
    words_clean = [w for w in words if w.lower() not in noise_words and not w.isdigit()]

    # Try 2-word combinations
    for i in range(len(words_clean) - 1):
        two_word = f"{words_clean[i]} {words_clean[i+1]}"
        if len(two_word) > 4:
            candidates.append(two_word)

    # Try 3-word combinations
    for i in range(len(words_clean) - 2):
        three_word = f"{words_clean[i]} {words_clean[i+1]} {words_clean[i+2]}"
        if len(three_word) > 6:
            candidates.append(three_word)

    # Deduplicate while preserving order
    seen = set()
    unique_candidates = []
    for candidate in candidates:
        candidate_lower = candidate.lower()
        if candidate_lower not in seen:
            seen.add(candidate_lower)
            unique_candidates.append(candidate)

    return unique_candidates


def extract_company_from_excel_content(file_content: bytes, max_rows: int = 10, max_cols: int = 10) -> List[str]:
    """
    Scan Excel file content for potential company names.

    Scans first N rows and columns looking for:
    - Company names in headers or metadata cells
    - Customer/client name fields
    - Legal entity names

    Args:
        file_content: Excel file as bytes
        max_rows: Maximum rows to scan (default: 10)
        max_cols: Maximum columns to scan (default: 10)

    Returns:
        List of potential company names found in content
    """
    candidates: Set[str] = set()

    try:
        wb = openpyxl.load_workbook(BytesIO(file_content), read_only=True, data_only=True)

        # Scan all sheets (usually just one for pay files)
        for sheet_name in wb.sheetnames[:3]:  # Limit to first 3 sheets
            ws = wb[sheet_name]

            # Scan first N rows and columns
            for row in ws.iter_rows(min_row=1, max_row=min(max_rows, ws.max_row),
                                   min_col=1, max_col=min(max_cols, ws.max_column),
                                   values_only=True):
                for cell_value in row:
                    if not cell_value:
                        continue

                    cell_str = str(cell_value).strip()

                    # Skip if too short or too long
                    if len(cell_str) < 3 or len(cell_str) > 100:
                        continue

                    # Skip if looks like a number or date
                    if cell_str.replace('.', '').replace(',', '').replace('-', '').replace('/', '').isdigit():
                        continue

                    # Look for company-like patterns
                    # Pattern 1: Contains "Ltd", "Limited", "PLC", "Inc", etc.
                    if re.search(r'\b(ltd|limited|plc|inc|corp|llc|gmbh)\b', cell_str, re.IGNORECASE):
                        candidates.add(cell_str)
                        continue

                    # Pattern 2: Has "Customer", "Client", "Company" label nearby
                    if re.search(r'(customer|client|company|supplier|vendor):', cell_str, re.IGNORECASE):
                        # Extract the value after the label
                        match = re.search(r'(customer|client|company|supplier|vendor):\s*(.+)', cell_str, re.IGNORECASE)
                        if match:
                            company_name = match.group(2).strip()
                            if len(company_name) > 2:
                                candidates.add(company_name)

                    # Pattern 3: Looks like a company name (2-4 words, capitalized)
                    words = cell_str.split()
                    if 2 <= len(words) <= 5:
                        # Check if mostly capitalized (proper nouns)
                        capitalized_count = sum(1 for w in words if w and w[0].isupper())
                        if capitalized_count >= len(words) * 0.6:  # 60% capitalized
                            candidates.add(cell_str)

        wb.close()

        return list(candidates)

    except Exception as e:
        print(f"Error scanning Excel content for company names: {str(e)}")
        return []


def match_file_to_canonical_groups(
    filename: str,
    file_content: Optional[bytes] = None,
    min_score: float = 0.75
) -> Dict:
    """
    Match a file to canonical customer groups using filename and content analysis.

    Process:
    1. Extract potential company names from filename
    2. If file_content provided, extract from Excel content
    3. Match all candidates against canonical groups
    4. Return best matches with confidence scores

    Args:
        filename: Original filename
        file_content: Optional Excel file content as bytes
        min_score: Minimum similarity score for matches (default: 0.75)

    Returns:
        Dict with matching results:
        {
            'best_match': {
                'canonical_id': 'CG-0001',
                'group_name': 'Nasstar Group',
                'confidence': 0.95,
                'matched_via': 'filename',
                'matched_text': 'Nasstar'
            },
            'filename_candidates': ['Nasstar', 'Payfile'],
            'content_candidates': ['Nasstar (UK) Limited'],
            'all_matches': [<list of all matches>],
            'match_found': True
        }
    """
    result = {
        'filename_candidates': [],
        'content_candidates': [],
        'all_matches': [],
        'best_match': None,
        'match_found': False
    }

    # Extract candidates from filename
    filename_candidates = extract_company_from_filename(filename)
    result['filename_candidates'] = filename_candidates

    # Extract candidates from content if provided
    content_candidates = []
    if file_content:
        content_candidates = extract_company_from_excel_content(file_content)
        result['content_candidates'] = content_candidates

    # Combine all candidates
    all_candidates = filename_candidates + content_candidates

    if not all_candidates:
        return result

    # Match each candidate against canonical groups
    all_matches = []

    for candidate in all_candidates:
        matches = canonical_groups_manager.match_entity(
            entity_name=candidate,
            active_only=True,
            min_score=min_score
        )

        # Tag matches with source
        source = 'filename' if candidate in filename_candidates else 'content'
        for match in matches:
            match['matched_via'] = source
            match['confidence'] = match['score']  # Alias for clarity
            all_matches.append(match)

    # Sort by confidence/score
    all_matches.sort(key=lambda x: x['confidence'], reverse=True)

    result['all_matches'] = all_matches

    # Set best match
    if all_matches:
        result['best_match'] = all_matches[0]
        result['match_found'] = True

    return result


def format_match_summary(match_result: Dict) -> str:
    """
    Create human-readable summary of matching results.

    Args:
        match_result: Result from match_file_to_canonical_groups()

    Returns:
        Formatted summary string
    """
    if not match_result['match_found']:
        return "No canonical group match found"

    best = match_result['best_match']
    confidence_pct = int(best['confidence'] * 100)

    return (
        f"{best['group_name']} "
        f"(matched '{best['matched_text']}' via {best['matched_via']}, "
        f"{confidence_pct}% confidence)"
    )
