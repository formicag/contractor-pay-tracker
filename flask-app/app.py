"""
Contractor Pay Tracker - Flask Web Application

Main Flask application with smart port selection
"""

import os
import socket
from datetime import datetime
import uuid
import json
import logging
import webbrowser
import threading
import time

from flask import Flask, render_template, request, jsonify, redirect, url_for, send_file
from dotenv import load_dotenv
import boto3
from boto3.dynamodb.conditions import Key
from botocore.exceptions import ClientError
from decimal import Decimal
from logging_config import setup_logging, log_request, log_response, log_errors
import openpyxl
import io

# Load environment variables
load_dotenv()

app = Flask(__name__)

# Setup comprehensive logging
setup_logging(app)
log_request(app)
log_response(app)
log_errors(app)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-secret-key-change-in-production')

# AWS Configuration
AWS_REGION = os.environ.get('AWS_REGION', 'eu-west-2')
AWS_PROFILE = os.environ.get('AWS_PROFILE', 'AdministratorAccess-016164185850')
S3_BUCKET = os.environ.get('S3_BUCKET_NAME', '')
DYNAMODB_TABLE = os.environ.get('DYNAMODB_TABLE_NAME', '')
PAYFILES_TABLE = os.environ.get('PAYFILES_TABLE_NAME', 'contractor-pay-files-development')
CONTRACTORS_TABLE = os.environ.get('CONTRACTORS_TABLE_NAME', 'contractor-pay-contractors-development')
PERMSTAFF_TABLE = os.environ.get('PERMSTAFF_TABLE_NAME', 'contractor-pay-permstaff-development')
UMBRELLAS_TABLE = os.environ.get('UMBRELLAS_TABLE_NAME', 'contractor-pay-umbrellas-development')
FILE_PROCESSOR_ARN = os.environ.get('FILE_PROCESSOR_ARN', '')

# Initialize AWS clients
session = boto3.Session(profile_name=AWS_PROFILE, region_name=AWS_REGION)
s3_client = session.client('s3')
dynamodb = session.resource('dynamodb')
lambda_client = session.client('lambda')
table = dynamodb.Table(DYNAMODB_TABLE) if DYNAMODB_TABLE else None
payfiles_table = dynamodb.Table(PAYFILES_TABLE) if PAYFILES_TABLE else None
contractors_table = dynamodb.Table(CONTRACTORS_TABLE)
permstaff_table = dynamodb.Table(PERMSTAFF_TABLE)
umbrellas_table = dynamodb.Table(UMBRELLAS_TABLE)

print("=" * 80)
print("🔧 AWS TABLES INITIALIZED")
print("=" * 80)
print(f"✅ Main table: {DYNAMODB_TABLE}")
print(f"✅ Pay files table: {PAYFILES_TABLE}")
print(f"✅ Contractors table: {CONTRACTORS_TABLE}")
print(f"✅ Perm staff table: {PERMSTAFF_TABLE}")
print(f"⛔ Umbrellas table (READ-ONLY - DO NOT MODIFY): {UMBRELLAS_TABLE}")
print("=" * 80)
print("⚠️  WARNING: contractor-pay-umbrellas-development is READ-ONLY")
print("⚠️  NO write operations allowed on umbrella companies table")
print("⚠️  See: DO-NOT-TOUCH-UMBRELLAS.md")
print("=" * 80)

# Register file management endpoints
from file_management_endpoints import register_file_endpoints
if payfiles_table:
    register_file_endpoints(app, payfiles_table, s3_client)


def find_available_port(start_port=5555, max_attempts=100):
    """
    Find an available port starting from start_port

    Args:
        start_port: Starting port number (default: 5555)
        max_attempts: Maximum number of ports to try

    Returns:
        int: Available port number
    """
    for port in range(start_port, start_port + max_attempts):
        try:
            # Try to bind to the port
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(1)
            result = sock.connect_ex(('127.0.0.1', port))
            sock.close()

            if result != 0:  # Port is available
                return port
        except socket.error:
            continue

    # Fallback to a random high port if all attempts fail
    return start_port + max_attempts


@app.route('/')
def index():
    """Home page - redirect to upload page"""
    return redirect(url_for('upload'))


@app.route('/upload')
def upload():
    """File upload page"""
    return render_template('upload.html')


@app.route('/files')
def files():
    """Files management dashboard"""
    return render_template('files.html')


@app.route('/file/<file_id>')
def file_detail(file_id):
    """File detail page with validation errors"""
    return render_template('file_detail.html', file_id=file_id)


@app.route('/contractors')
def contractors():
    """Contractors management page"""
    return render_template('contractors.html')


@app.route('/perm-staff')
def perm_staff():
    """Permanent staff management page"""
    return render_template('perm_staff.html')


################################################################################
# ⛔ LOCKED CODE - DO NOT MODIFY ⛔
# Umbrella Dashboard Route - READ-ONLY FOREVER
# See: DO-NOT-TOUCH-UMBRELLA-DASHBOARD.md
################################################################################
@app.route('/umbrella-dashboard')
def umbrella_dashboard():
    """Umbrella company reporting dashboard"""
    return render_template('umbrella_dashboard.html')
################################################################################
# ⛔ END LOCKED CODE ⛔
################################################################################


@app.route('/debug')
def debug():
    """Debugging interface for comparing Excel files with database records"""
    return render_template('debug.html')


@app.route('/ops/debug-files')
def debug_files():
    """Debug files page - Three-panel view for operations"""
    return render_template('debug_files.html')


@app.route('/api/upload', methods=['POST'])
def api_upload():
    """
    API endpoint for file upload

    Uploads contractor pay file to S3 and creates metadata in DynamoDB

    Expected: multipart/form-data with file
    Returns: JSON with file_id and status
    """
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400

        file = request.files['file']

        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400

        if not file.filename.endswith('.xlsx'):
            return jsonify({'error': 'Only .xlsx files are supported'}), 400

        # Generate unique file ID (UUID)
        import time
        file_id = str(uuid.uuid4())
        timestamp = int(time.time())
        s3_key = f"production/{file.filename}"

        # Read file content
        file_content = file.read()
        file_size = len(file_content)

        # Upload to S3
        s3_client.put_object(
            Bucket=S3_BUCKET,
            Key=s3_key,
            Body=file_content,
            ContentType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        # Extract umbrella company from filename
        umbrella_company = file.filename.split(' ')[0] if ' ' in file.filename else 'Unknown'

        # Extract contractor names from Excel file
        from excel_processor import extract_contractor_names, get_contractor_summary
        contractor_names = extract_contractor_names(file_content)
        contractor_name = get_contractor_summary(contractor_names)

        app.logger.info(f"Extracted {len(contractor_names)} contractors from {file.filename}: {contractor_names}")

        # Match file to canonical customer groups
        from canonical_matcher import match_file_to_canonical_groups, format_match_summary
        canonical_match = match_file_to_canonical_groups(
            filename=file.filename,
            file_content=file_content,
            min_score=0.75
        )

        # Extract match details for storage
        canonical_id = canonical_match['best_match']['canonical_id'] if canonical_match['match_found'] else None
        canonical_group_name = canonical_match['best_match']['group_name'] if canonical_match['match_found'] else None
        match_confidence = canonical_match['best_match']['confidence'] if canonical_match['match_found'] else None
        match_via = canonical_match['best_match']['matched_via'] if canonical_match['match_found'] else None
        match_summary = format_match_summary(canonical_match)

        app.logger.info(f"Canonical group match for {file.filename}: {match_summary}")
        if canonical_match['match_found']:
            app.logger.info(f"  - Filename candidates: {canonical_match['filename_candidates']}")
            app.logger.info(f"  - Content candidates: {canonical_match['content_candidates']}")
            app.logger.info(f"  - Total matches: {len(canonical_match['all_matches'])}")

        # Create File metadata record in DynamoDB using PascalCase schema
        from decimal import Decimal
        uploaded_at_iso = datetime.utcnow().isoformat() + 'Z'

        item_data = {
            # Primary Key fields
            'PK': f'FILE#{file_id}',
            'SK': 'METADATA',
            'EntityType': 'File',

            # File metadata (PascalCase)
            'FileID': file_id,
            'OriginalFilename': file.filename,
            'S3Key': s3_key,
            'S3Bucket': S3_BUCKET,
            'FileSizeBytes': file_size,
            'UmbrellaCode': umbrella_company.upper(),
            'umbrella_code': umbrella_company.upper(),  # Also store lowercase for compatibility
            'contractor_name': contractor_name,
            'contractor_names_list': contractor_names,
            'contractor_count': len(contractor_names),
            'UploadedAt': uploaded_at_iso,
            'Status': 'UPLOADED',
            'processing_status': 'completed' if contractor_names else 'failed',
            'Version': 1,
            'IsCurrentVersion': True,

            # GSI fields for querying
            'GSI3PK': 'FILES',
            'GSI3SK': uploaded_at_iso
        }

        # Add canonical group match data if found
        if canonical_match['match_found']:
            item_data['canonical_id'] = canonical_id
            item_data['canonical_group_name'] = canonical_group_name
            item_data['canonical_match_confidence'] = Decimal(str(match_confidence))
            item_data['canonical_match_via'] = match_via
            item_data['canonical_match_summary'] = match_summary

        payfiles_table.put_item(Item=item_data)

        app.logger.info(f"File uploaded successfully: {file_id} -> s3://{S3_BUCKET}/{s3_key}")

        response_data = {
            'message': 'File uploaded successfully',
            'file_id': file_id,
            'filename': file.filename,
            'size': file_size,
            'umbrella_company': umbrella_company,
            's3_key': s3_key,
            'status': 'uploaded',
            'contractor_count': len(contractor_names)
        }

        # Add canonical match info to response
        if canonical_match['match_found']:
            response_data['canonical_match'] = {
                'found': True,
                'canonical_id': canonical_id,
                'group_name': canonical_group_name,
                'confidence': match_confidence,
                'matched_via': match_via,
                'summary': match_summary
            }
        else:
            response_data['canonical_match'] = {
                'found': False,
                'message': 'No canonical group match found'
            }

        return jsonify(response_data)

    except Exception as e:
        app.logger.error(f"Error uploading file: {str(e)}")
        import traceback
        app.logger.error(traceback.format_exc())
        return jsonify({'error': 'Failed to upload file', 'message': str(e)}), 500


@app.route('/api/database-info', methods=['GET'])
def api_database_info():
    """
    API endpoint to get database information including ARN

    Returns: JSON with database metadata
    """
    try:
        # Get table ARN from DynamoDB
        table_arn = f"arn:aws:dynamodb:{AWS_REGION}:016164185850:table/{PAYFILES_TABLE}"

        return jsonify({
            'table_name': PAYFILES_TABLE,
            'table_arn': table_arn,
            'region': AWS_REGION,
            'bucket_name': S3_BUCKET
        })

    except Exception as e:
        app.logger.error(f"Error fetching database info: {str(e)}")
        return jsonify({'error': 'Failed to fetch database info', 'message': str(e)}), 500


@app.route('/api/files', methods=['GET'])
def api_files():
    """
    API endpoint to list files

    Returns: JSON with files list
    """
    try:
        # Query DynamoDB payfiles table for all uploaded files
        # Filter for EntityType = 'File' to get only file records
        response = payfiles_table.scan(
            FilterExpression='EntityType = :et',
            ExpressionAttributeValues={':et': 'File'}
        )

        files_list = []
        for item in response.get('Items', []):
            # Extract file info using the correct PascalCase attribute names
            file_id = item.get('FileID')

            # Parse uploaded_at timestamp
            uploaded_at_str = item.get('UploadedAt', '')
            try:
                from datetime import datetime
                uploaded_at = int(datetime.fromisoformat(uploaded_at_str.replace('Z', '+00:00')).timestamp())
            except:
                uploaded_at = 0

            file_data = {
                'file_id': file_id,
                'filename': item.get('OriginalFilename', 'Unknown'),
                'contractor_name': item.get('contractor_name', 'N/A'),
                'umbrella_company': item.get('UmbrellaCode', item.get('umbrella_code', 'Unknown')),
                'umbrella': item.get('UmbrellaCode', item.get('umbrella_code', 'Unknown')),
                'status': item.get('Status', 'UNKNOWN'),
                'processing_status': item.get('processing_status', item.get('Status', 'pending')),
                'uploaded_at': uploaded_at,
                'file_size': int(item.get('FileSizeBytes', 0)),
                's3_key': item.get('S3Key', '')
            }

            files_list.append(file_data)

        # Sort by upload date (newest first)
        files_list.sort(key=lambda x: x.get('uploaded_at', 0), reverse=True)

        return jsonify({'files': files_list, 'total': len(files_list)})

    except Exception as e:
        app.logger.error(f"Error fetching files: {str(e)}")
        import traceback
        app.logger.error(traceback.format_exc())
        return jsonify({'error': 'Failed to fetch files', 'message': str(e)}), 500

# Old broken endpoints removed - replaced by file_management_endpoints.py


@app.route('/api/files/<file_id>/navigation', methods=['GET'])
def api_file_navigation(file_id):
    """
    API endpoint to get navigation info (previous/next file IDs)

    Query Parameters:
        - status: Filter by status (optional)
        - umbrella_code: Filter by umbrella company code (optional)

    Returns: JSON with previous and next file IDs, current index, and total
    """
    try:
        # Get query parameters for filtering
        status_filter = request.args.get('status')
        umbrella_filter = request.args.get('umbrella_code')

        # Build filter expression
        filter_expr = 'EntityType = :et'
        expression_values = {':et': 'File'}
        expression_names = {}

        if status_filter:
            filter_expr += ' AND #status = :status'
            expression_values[':status'] = status_filter
            expression_names['#status'] = 'Status'

        if umbrella_filter:
            filter_expr += ' AND UmbrellaCode = :umbrella'
            expression_values[':umbrella'] = umbrella_filter

        # Query DynamoDB
        scan_kwargs = {
            'FilterExpression': filter_expr,
            'ExpressionAttributeValues': expression_values
        }

        if expression_names:
            scan_kwargs['ExpressionAttributeNames'] = expression_names

        response = table.scan(**scan_kwargs)

        # Build and sort file list (oldest first)
        files_list = []
        for item in response.get('Items', []):
            files_list.append({
                'file_id': item.get('FileID'),
                'uploaded_at': item.get('UploadedAt', item.get('CreatedAt', ''))
            })

        files_list.sort(key=lambda x: x.get('uploaded_at', ''))

        # Find current file index
        current_index = None
        for idx, file_item in enumerate(files_list):
            if file_item['file_id'] == file_id:
                current_index = idx
                break

        if current_index is None:
            return jsonify({'error': 'File not found'}), 404

        # Get previous and next file IDs
        previous_id = files_list[current_index - 1]['file_id'] if current_index > 0 else None
        next_id = files_list[current_index + 1]['file_id'] if current_index < len(files_list) - 1 else None

        return jsonify({
            'previous': previous_id,
            'next': next_id,
            'current_index': current_index + 1,  # 1-indexed for display
            'total': len(files_list)
        })

    except Exception as e:
        app.logger.error(f"Error getting file navigation: {str(e)}")
        return jsonify({'error': 'Failed to get navigation info', 'message': str(e)}), 500


@app.route('/api/files/<file_id>/xls', methods=['GET'])
def api_file_xls(file_id):
    """
    API endpoint to fetch Excel file from S3 and return as JSON

    Returns: JSON with Excel data organized by sheets
    """
    try:
        # Get file metadata
        file_response = table.get_item(
            Key={'PK': f'FILE#{file_id}', 'SK': 'METADATA'}
        )

        if 'Item' not in file_response:
            return jsonify({'error': 'File not found'}), 404

        file_metadata = file_response['Item']
        s3_bucket = file_metadata.get('S3Bucket', S3_BUCKET)
        s3_key = file_metadata.get('S3Key')

        if not s3_key:
            return jsonify({'error': 'S3 key not found in file metadata'}), 404

        # Download file from S3
        file_obj = io.BytesIO()
        s3_client.download_fileobj(s3_bucket, s3_key, file_obj)
        file_obj.seek(0)

        # Load workbook
        workbook = openpyxl.load_workbook(file_obj, data_only=True)

        # Convert to JSON structure
        sheets_data = {}
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_data = []

            # Get all rows
            for row in sheet.iter_rows(values_only=True):
                # Convert Decimal to float for JSON serialization
                row_data = []
                for cell in row:
                    if isinstance(cell, Decimal):
                        row_data.append(float(cell))
                    elif cell is None:
                        row_data.append('')
                    else:
                        row_data.append(cell)
                sheet_data.append(row_data)

            sheets_data[sheet_name] = sheet_data

        return jsonify({
            'file_id': file_id,
            'filename': file_metadata.get('OriginalFilename', 'Unknown'),
            'sheets': sheets_data
        })

    except ClientError as e:
        app.logger.error(f"S3 error fetching file {file_id}: {str(e)}")
        return jsonify({'error': 'Failed to fetch file from S3', 'message': str(e)}), 500
    except Exception as e:
        app.logger.error(f"Error fetching Excel file {file_id}: {str(e)}")
        return jsonify({'error': 'Failed to process Excel file', 'message': str(e)}), 500


@app.route('/api/files/<file_id>/database', methods=['GET'])
def api_file_database(file_id):
    """
    API endpoint to get all database records for a file

    Returns: JSON with file metadata, pay records, and errors
    """
    try:
        # Get file metadata
        file_response = table.get_item(
            Key={'PK': f'FILE#{file_id}', 'SK': 'METADATA'}
        )

        if 'Item' not in file_response:
            return jsonify({'error': 'File not found'}), 404

        file_metadata = file_response['Item']

        # Query all items for this file
        response = table.query(
            KeyConditionExpression=Key('PK').eq(f'FILE#{file_id}')
        )

        # Organize data by type
        pay_records = []
        validation_errors = []
        validation_warnings = []

        for item in response.get('Items', []):
            entity_type = item.get('EntityType', '')

            if entity_type == 'PayRecord':
                # Convert Decimal to float
                record_data = {k: (float(v) if isinstance(v, Decimal) else v) for k, v in item.items()}
                pay_records.append(record_data)
            elif entity_type == 'ValidationError':
                validation_errors.append(item)
            elif entity_type == 'ValidationWarning':
                validation_warnings.append(item)

        # Sort pay records by SK (record number)
        pay_records.sort(key=lambda x: x.get('SK', ''))

        return jsonify({
            'file_metadata': {k: (float(v) if isinstance(v, Decimal) else v) for k, v in file_metadata.items()},
            'pay_records': pay_records,
            'pay_records_count': len(pay_records),
            'validation_errors': validation_errors,
            'validation_warnings': validation_warnings
        })

    except Exception as e:
        app.logger.error(f"Error fetching database data for file {file_id}: {str(e)}")
        return jsonify({'error': 'Failed to fetch database data', 'message': str(e)}), 500


@app.route('/api/files/<file_id>/validation', methods=['GET'])
def api_file_validation(file_id):
    """
    API endpoint to get validation results for a file

    Returns: JSON with validation snapshot and all errors/warnings
    """
    try:
        # Query for ValidationSnapshot (most recent)
        snapshot_response = table.query(
            KeyConditionExpression=Key('PK').eq(f'FILE#{file_id}') & Key('SK').begins_with('VALIDATION#'),
            ScanIndexForward=False,  # Get most recent first
            Limit=1
        )

        # Query validation errors for detailed info
        errors_response = table.query(
            KeyConditionExpression=Key('PK').eq(f'FILE#{file_id}') & Key('SK').begins_with('ERROR#')
        )

        # Query validation warnings
        warnings_response = table.query(
            KeyConditionExpression=Key('PK').eq(f'FILE#{file_id}') & Key('SK').begins_with('WARNING#')
        )

        # Get file metadata for status
        file_response = table.get_item(
            Key={'PK': f'FILE#{file_id}', 'SK': 'METADATA'}
        )

        file_status = 'UNKNOWN'
        if 'Item' in file_response:
            file_status = file_response['Item'].get('Status', 'UNKNOWN')

        # Use ValidationSnapshot if available, otherwise build from errors/warnings
        validation_rules = []
        snapshot_data = None

        if snapshot_response.get('Items'):
            snapshot = snapshot_response['Items'][0]
            snapshot_data = {
                'status': snapshot.get('Status'),
                'validated_at': snapshot.get('ValidatedAt'),
                'total_records': snapshot.get('TotalRecords'),
                'valid_records': snapshot.get('ValidRecords'),
                'rejected_records': snapshot.get('RejectedRecords', 0),
                'total_checks': snapshot.get('TotalChecks', 0),
                'passed_checks': snapshot.get('PassedChecks', 0),
                'critical_failures': snapshot.get('CriticalFailures', 0),
                'warnings': snapshot.get('Warnings', 0),
                # Legacy fields for backwards compatibility
                'total_rules': snapshot.get('TotalRules', 0),
                'passed_rules': snapshot.get('PassedRules', 0),
                'failed_rules': snapshot.get('FailedRules', 0),
                'error_count': snapshot.get('ErrorCount', 0),
                'warning_count': snapshot.get('WarningCount', 0)
            }

            # Use new enterprise validation checks if available
            all_validation_checks = snapshot.get('AllValidationChecks', [])

            if all_validation_checks:
                # Return all validation checks directly (enterprise format)
                validation_rules = all_validation_checks
            else:
                # Fallback to old format for backwards compatibility
                validation_results = snapshot.get('ValidationResults', {})

                for error_type, rule_data in validation_results.items():
                    validation_rules.append({
                        'name': error_type.replace('_', ' ').title(),
                        'key': error_type,
                        'passed': rule_data.get('passed', True),
                        'severity': rule_data.get('severity', 'UNKNOWN'),
                        'is_warning': rule_data.get('severity') == 'WARNING',
                        'messages': rule_data.get('messages', []),
                        'affected_count': len(rule_data.get('affected_records', []))
                    })

        else:
            # Fallback: build from individual error/warning records
            rule_map = {}

            for error in errors_response.get('Items', []):
                error_type = error.get('ErrorType', '')
                if error_type not in rule_map:
                    rule_map[error_type] = {
                        'name': error_type.replace('_', ' ').title(),
                        'key': error_type,
                        'passed': False,
                        'severity': error.get('Severity', 'CRITICAL'),
                        'is_warning': False,
                        'messages': [],
                        'errors': []
                    }

                rule_map[error_type]['errors'].append({
                    'row': error.get('RowNumber'),
                    'employee_id': error.get('EmployeeID'),
                    'contractor_name': error.get('ContractorName'),
                    'message': error.get('ErrorMessage')
                })
                if error.get('ErrorMessage'):
                    rule_map[error_type]['messages'].append(error.get('ErrorMessage'))

            for warning in warnings_response.get('Items', []):
                warning_type = warning.get('WarningType', '')
                if warning_type not in rule_map:
                    rule_map[warning_type] = {
                        'name': warning_type.replace('_', ' ').title(),
                        'key': warning_type,
                        'passed': True,
                        'severity': 'WARNING',
                        'is_warning': True,
                        'messages': [],
                        'errors': []
                    }

                rule_map[warning_type]['errors'].append({
                    'row': warning.get('RowNumber'),
                    'message': warning.get('WarningMessage'),
                    'auto_resolved': warning.get('AutoResolved', False)
                })
                if warning.get('WarningMessage'):
                    rule_map[warning_type]['messages'].append(warning.get('WarningMessage'))

            validation_rules = list(rule_map.values())

        return jsonify({
            'file_id': file_id,
            'file_status': file_status,
            'snapshot': snapshot_data,
            'has_errors': len(errors_response.get('Items', [])) > 0,
            'has_warnings': len(warnings_response.get('Items', [])) > 0,
            'total_errors': len(errors_response.get('Items', [])),
            'total_warnings': len(warnings_response.get('Items', [])),
            'validation_rules': validation_rules
        })

    except Exception as e:
        app.logger.error(f"Error fetching validation data for file {file_id}: {str(e)}")
        return jsonify({'error': 'Failed to fetch validation data', 'message': str(e)}), 500


@app.route('/api/contractors/lock-status', methods=['GET'])
def api_contractors_lock_status():
    """
    API endpoint to get contractor data lock status

    Returns: JSON with lock status and protection information
    """
    from contractor_lock import get_contractor_lock_status
    return jsonify(get_contractor_lock_status())


@app.route('/api/contractors', methods=['GET'])
def api_contractors():
    """
    API endpoint to list contractors (READ-ONLY)
    🔒 This data is protected - modifications require password

    Returns: JSON with contractors list
    """
    try:
        # Query DynamoDB for Contractor entities from contractor-pay-development table
        response = payfiles_table.scan(
            FilterExpression='EntityType = :et',
            ExpressionAttributeValues={
                ':et': 'Contractor'
            }
        )

        contractors_list = []
        for item in response.get('Items', []):
            # Extract data using PascalCase schema
            contractor_id = item.get('ContractorID', '')
            email = item.get('Email', item.get('ResourceEmail', ''))
            first_name = item.get('FirstName', '')
            last_name = item.get('LastName', '')

            # Convert Decimal to float for rates
            sell_rate = float(item.get('SellRate', 0))

            contractor_data = {
                'contractor_id': contractor_id,
                'email': email,
                'first_name': first_name,
                'last_name': last_name,
                'full_name': f"{first_name} {last_name}".strip(),
                'job_title': item.get('JobTitle', 'N/A'),
                'umbrella_company': item.get('UmbrellaCompany', 'N/A'),
                'sell_rate': sell_rate,
                'is_active': item.get('IsActive', True),
                'created_at': item.get('CreatedAt', ''),
                'engagement_type': item.get('EngagementType', 'N/A'),
                'customer': item.get('Customer', 'N/A')
            }
            contractors_list.append(contractor_data)

        # Sort by last name
        contractors_list.sort(key=lambda x: x.get('last_name', ''))

        return jsonify({'contractors': contractors_list})

    except Exception as e:
        app.logger.error(f"Error fetching contractors: {str(e)}")
        import traceback
        app.logger.error(traceback.format_exc())
        return jsonify({'error': 'Failed to fetch contractors', 'message': str(e)}), 500


@app.route('/api/perm-staff/lock-status', methods=['GET'])
def api_perm_staff_lock_status():
    """
    API endpoint to get permanent staff data lock status

    Returns: JSON with lock status and protection information
    """
    from perm_staff_lock import get_lock_status
    return jsonify(get_lock_status())


@app.route('/api/perm-staff', methods=['GET'])
def api_perm_staff():
    """
    API endpoint to list permanent staff (READ-ONLY)
    🔒 This data is protected - modifications require password

    IMPORTANT: Permanent staff are stored as EntityType = 'PermanentStaff'
    They are completely separate from contractors.

    Returns: JSON with permanent staff list
    """
    try:
        app.logger.info("=" * 80)
        app.logger.info("🔍 PERMANENT STAFF API - DETAILED VERBOSE LOGGING")
        app.logger.info("=" * 80)

        # STEP 1: Log table information
        app.logger.info("STEP 1: Identifying table to scan")
        app.logger.info(f"   Table name: {PERMSTAFF_TABLE}")
        app.logger.info(f"   Table object: {permstaff_table}")
        app.logger.info(f"   EntityType filter: PermanentStaff")

        # STEP 2: Scan DynamoDB for PermanentStaff entities
        app.logger.info("=" * 80)
        app.logger.info("STEP 2: Scanning DynamoDB for permanent staff")
        app.logger.info("=" * 80)

        response = permstaff_table.scan(
            FilterExpression='EntityType = :et',
            ExpressionAttributeValues={
                ':et': 'PermanentStaff'
            }
        )

        app.logger.info(f"✅ Scan complete")
        app.logger.info(f"   Items returned: {len(response.get('Items', []))}")
        app.logger.info(f"   Scanned count: {response.get('ScannedCount', 0)}")

        # STEP 3: Process each staff member
        app.logger.info("=" * 80)
        app.logger.info("STEP 3: Processing staff members into list")
        app.logger.info("=" * 80)

        staff_list = []
        for idx, item in enumerate(response.get('Items', [])):
            app.logger.info(f"Processing staff member {idx + 1}/{len(response.get('Items', []))}")
            app.logger.info(f"   Raw item keys: {list(item.keys())}")

            # STEP 3a: Extract identification fields
            staff_id = item.get('StaffID', '')
            employee_id = item.get('EmployeeID', '')

            app.logger.info(f"   StaffID: {staff_id}")

            # STEP 3b: Extract name fields
            first_name = item.get('FirstName', '')
            last_name = item.get('LastName', '')
            full_name = f"{first_name} {last_name}".strip()

            app.logger.info(f"   Name: {full_name}")

            # STEP 3c: Extract role/job fields
            job_title = item.get('JobTitle', 'N/A')
            customer = item.get('Customer', 'N/A')
            sow_service_name = item.get('SowServiceName', 'N/A')

            app.logger.info(f"   JobTitle: {job_title}")
            app.logger.info(f"   Customer: {customer}")
            app.logger.info(f"   SowServiceName: {sow_service_name}")

            # STEP 3d: Extract rate (convert Decimal to float)
            # Use SellRate only
            rate_decimal = item.get('SellRate', 0)
            sell_rate = float(rate_decimal) if rate_decimal else 0.0

            app.logger.info(f"   SellRate: £{sell_rate}")

            # STEP 3e: Extract other fields
            is_active = item.get('IsActive', True)
            created_at = item.get('CreatedAt', '')

            # STEP 3f: Build staff data object - ALL FIELDS
            staff_data = {
                'pk': item.get('PK', ''),
                'sk': item.get('SK', ''),
                'staff_id': staff_id,
                'employee_id': employee_id,
                'first_name': first_name,
                'last_name': last_name,
                'full_name': full_name,
                'job_title': job_title,
                'customer': customer,
                'engagement_type': 'PAYE',
                'sow_service_name': sow_service_name,
                'sell_rate': sell_rate,
                'currency': item.get('Currency', 'GBP'),
                'is_active': is_active,
                'entity_type': item.get('EntityType', ''),
                'created_at': created_at,
                'updated_at': item.get('UpdatedAt', ''),
                'gsi1_pk': item.get('GSI1PK', ''),
                'gsi1_sk': item.get('GSI1SK', ''),
                'gsi2_pk': item.get('GSI2PK', ''),
                'gsi2_sk': item.get('GSI2SK', '')
            }

            staff_list.append(staff_data)
            app.logger.info(f"   ✅ Added {full_name} to staff list")

        # STEP 4: Sort staff list
        app.logger.info("=" * 80)
        app.logger.info("STEP 4: Sorting staff list by last name")
        app.logger.info("=" * 80)

        staff_list.sort(key=lambda x: x.get('last_name', ''))

        app.logger.info(f"✅ Sorted {len(staff_list)} staff members")

        # STEP 5: Build response
        app.logger.info("=" * 80)
        app.logger.info("STEP 5: Building JSON response")
        app.logger.info("=" * 80)

        response_data = {'staff': staff_list}

        app.logger.info(f"   Staff count in response: {len(staff_list)}")
        for staff in staff_list:
            app.logger.info(f"      - {staff['full_name']} - {staff['job_title']} - {staff['customer']} - £{staff['sell_rate']}")

        app.logger.info("=" * 80)
        app.logger.info("✅ API RESPONSE READY - Returning to client")
        app.logger.info(f"   Response contains {len(staff_list)} permanent staff members")
        app.logger.info("=" * 80)

        return jsonify(response_data)

    except Exception as e:
        app.logger.error("=" * 80)
        app.logger.error("❌ ERROR IN PERMANENT STAFF API")
        app.logger.error("=" * 80)
        app.logger.error(f"Error message: {str(e)}")
        app.logger.error(f"Error type: {type(e).__name__}")
        import traceback
        app.logger.error("Full traceback:")
        app.logger.error(traceback.format_exc())
        app.logger.error("=" * 80)
        return jsonify({'error': 'Failed to fetch permanent staff', 'message': str(e)}), 500


@app.route('/api/perm-staff-lambda', methods=['GET'])
def api_perm_staff_lambda():
    """
    API endpoint to fetch permanent staff via Lambda function
    🚀 Invokes dedicated Lambda that reads directly from DynamoDB

    This ensures we're getting data from the CORRECT table with CORRECT fields
    """
    import sys
    import os

    # Track API call start
    from debug_tracker import track_api_call
    start_time = time.time()

    try:
        app.logger.info("=" * 80)
        app.logger.info("🚀 INVOKING PERM STAFF READER LAMBDA")
        app.logger.info("=" * 80)

        # Import and run Lambda locally
        sys.path.insert(0, '/Users/gianlucaformica/Projects/contractor-pay-tracker/backend/functions/perm_staff_reader')
        import app as lambda_app

        # Invoke Lambda handler
        app.logger.info("Calling Lambda handler...")
        lambda_response = lambda_app.lambda_handler({}, {})

        app.logger.info(f"✅ Lambda returned status: {lambda_response['statusCode']}")

        # Parse response
        response_body = json.loads(lambda_response['body'])

        # Track successful API call
        response_time = (time.time() - start_time) * 1000
        track_api_call('/api/perm-staff-lambda', 'GET', 200, response_time, None, {'count': response_body.get('count')})

        app.logger.info(f"✅ Returning {response_body.get('count')} staff members from Lambda")
        app.logger.info("=" * 80)

        return jsonify(response_body)

    except Exception as e:
        app.logger.error("❌ ERROR INVOKING LAMBDA")
        app.logger.error(f"Error: {str(e)}")
        import traceback
        app.logger.error(traceback.format_exc())

        # Track failed API call
        response_time = (time.time() - start_time) * 1000
        track_api_call('/api/perm-staff-lambda', 'GET', 500, response_time, None, {'error': str(e)})

        from debug_tracker import track_error
        track_error('LambdaInvocationError', str(e), traceback.format_exc())

        return jsonify({'success': False, 'error': 'Failed to invoke Lambda', 'message': str(e)}), 500


@app.route('/api/debug/events', methods=['GET'])
def api_debug_events():
    """
    API endpoint to get debug events
    Returns recent debug events for the debug panel
    """
    from debug_tracker import debug_tracker

    since = request.args.get('since', type=int)
    limit = request.args.get('limit', 100, type=int)

    events = debug_tracker.get_events(since=since, limit=limit)

    return jsonify({
        'events': events,
        'count': len(events)
    })


@app.route('/api/debug/stream')
def api_debug_stream():
    """
    Server-Sent Events endpoint for real-time debug event streaming
    """
    from flask import Response, stream_with_context
    import queue
    import time
    from debug_tracker import debug_tracker

    def event_stream():
        """Generator for SSE stream"""
        q = queue.Queue()
        debug_tracker.subscribe(q)

        try:
            # Send initial connection message
            yield f"data: {json.dumps({'type': 'connected', 'message': 'Debug stream connected'})}\n\n"

            while True:
                try:
                    # Wait for new event (with timeout to send heartbeat)
                    event = q.get(timeout=30)
                    yield f"data: {json.dumps(event)}\n\n"
                except queue.Empty:
                    # Send heartbeat to keep connection alive
                    yield f": heartbeat\n\n"

        except GeneratorExit:
            debug_tracker.unsubscribe(q)

    return Response(
        stream_with_context(event_stream()),
        mimetype='text/event-stream',
        headers={
            'Cache-Control': 'no-cache',
            'X-Accel-Buffering': 'no'
        }
    )


@app.route('/api/debug/track', methods=['POST'])
def api_debug_track():
    """
    API endpoint to track user actions from frontend
    """
    from debug_tracker import track_user_action

    data = request.get_json()
    action = data.get('action', 'unknown')
    details = data.get('details', {})

    track_user_action(action, details)

    return jsonify({'success': True})


# ============================================================================
# CANONICAL CUSTOMER GROUPS API ENDPOINTS
# ============================================================================

@app.route('/api/canonical-groups', methods=['GET'])
def api_get_canonical_groups():
    """
    API endpoint to list all canonical customer groups

    Returns: JSON with list of all groups
    """
    try:
        from canonical_groups_manager import canonical_groups_manager

        groups = canonical_groups_manager.get_all_groups()

        return jsonify({
            'success': True,
            'groups': groups,
            'total': len(groups)
        })

    except Exception as e:
        app.logger.error(f"Error fetching canonical groups: {str(e)}")
        return jsonify({
            'success': False,
            'error': 'Failed to fetch canonical groups',
            'message': str(e)
        }), 500


@app.route('/api/canonical-groups/<canonical_id>', methods=['GET'])
def api_get_canonical_group(canonical_id):
    """
    API endpoint to get a specific canonical customer group

    Args:
        canonical_id: The canonical group ID (e.g., CG-0001)

    Returns: JSON with group details
    """
    try:
        from canonical_groups_manager import canonical_groups_manager

        group = canonical_groups_manager.get_group(canonical_id)

        if not group:
            return jsonify({
                'success': False,
                'error': 'Group not found',
                'canonical_id': canonical_id
            }), 404

        return jsonify({
            'success': True,
            'group': group
        })

    except Exception as e:
        app.logger.error(f"Error fetching canonical group {canonical_id}: {str(e)}")
        return jsonify({
            'success': False,
            'error': 'Failed to fetch canonical group',
            'message': str(e)
        }), 500


@app.route('/api/canonical-groups', methods=['POST'])
def api_create_canonical_group():
    """
    API endpoint to create a new canonical customer group
    🔒 LOCKED - Requires password "luca" to modify

    Request body:
        {
            "canonical_id": "CG-0003",
            "group_name": "Example Group",
            "legal_entity": "Example Ltd",
            "company_no": "12345678",
            "unlock_password": "luca"
        }

    Returns: JSON with success status and created group
    """
    try:
        from canonical_groups_manager import canonical_groups_manager

        data = request.get_json()

        # Extract required fields
        canonical_id = data.get('canonical_id')
        group_name = data.get('group_name')
        legal_entity = data.get('legal_entity')
        company_no = data.get('company_no')
        unlock_password = data.get('unlock_password') or request.headers.get('X-Unlock-Password')

        # Validate required fields
        if not all([canonical_id, group_name, legal_entity, company_no]):
            return jsonify({
                'success': False,
                'error': 'Missing required fields',
                'required': ['canonical_id', 'group_name', 'legal_entity', 'company_no']
            }), 400

        # Create group
        result = canonical_groups_manager.create_group(
            canonical_id=canonical_id,
            group_name=group_name,
            legal_entity=legal_entity,
            company_no=company_no,
            unlock_password=unlock_password,
            created_by='api'
        )

        if result['success']:
            return jsonify(result), 201
        else:
            status_code = 403 if 'password' in result.get('error', '').lower() else 400
            return jsonify(result), status_code

    except Exception as e:
        app.logger.error(f"Error creating canonical group: {str(e)}")
        return jsonify({
            'success': False,
            'error': 'Failed to create canonical group',
            'message': str(e)
        }), 500


@app.route('/api/canonical-groups/<canonical_id>/aliases', methods=['POST'])
def api_add_canonical_alias(canonical_id):
    """
    API endpoint to add an alias to a canonical customer group
    🔒 LOCKED - Requires password "luca" to modify

    Args:
        canonical_id: The canonical group ID

    Request body:
        {
            "alias": "Example Brand",
            "alias_type": "Brand",
            "start_date": "2020-01-01",
            "end_date": null,
            "notes": "Optional notes",
            "unlock_password": "luca"
        }

    Returns: JSON with success status and added alias
    """
    try:
        from canonical_groups_manager import canonical_groups_manager

        data = request.get_json()

        # Extract fields
        alias = data.get('alias')
        alias_type = data.get('alias_type')
        start_date = data.get('start_date')
        end_date = data.get('end_date')
        notes = data.get('notes', '')
        unlock_password = data.get('unlock_password') or request.headers.get('X-Unlock-Password')

        # Validate required fields
        if not all([alias, alias_type]):
            return jsonify({
                'success': False,
                'error': 'Missing required fields',
                'required': ['alias', 'alias_type']
            }), 400

        # Add alias
        result = canonical_groups_manager.add_alias(
            canonical_id=canonical_id,
            alias=alias,
            alias_type=alias_type,
            unlock_password=unlock_password,
            start_date=start_date,
            end_date=end_date,
            notes=notes,
            modified_by='api'
        )

        if result['success']:
            return jsonify(result), 201
        else:
            status_code = 403 if 'password' in result.get('error', '').lower() else 400
            return jsonify(result), status_code

    except Exception as e:
        app.logger.error(f"Error adding alias to group {canonical_id}: {str(e)}")
        return jsonify({
            'success': False,
            'error': 'Failed to add alias',
            'message': str(e)
        }), 500


@app.route('/api/canonical-groups/<canonical_id>', methods=['DELETE'])
def api_delete_canonical_group(canonical_id):
    """
    API endpoint to delete a canonical customer group
    🔒 LOCKED - Requires password "luca" to modify

    Args:
        canonical_id: The canonical group ID

    Request body or header:
        {
            "unlock_password": "luca"
        }
        OR
        X-Unlock-Password: luca

    Returns: JSON with success status
    """
    try:
        from canonical_groups_manager import canonical_groups_manager

        data = request.get_json() if request.is_json else {}
        unlock_password = data.get('unlock_password') or request.headers.get('X-Unlock-Password')

        # Delete group
        result = canonical_groups_manager.delete_group(
            canonical_id=canonical_id,
            unlock_password=unlock_password,
            deleted_by='api'
        )

        if result['success']:
            return jsonify(result), 200
        else:
            status_code = 403 if 'password' in result.get('error', '').lower() else 400
            return jsonify(result), status_code

    except Exception as e:
        app.logger.error(f"Error deleting canonical group {canonical_id}: {str(e)}")
        return jsonify({
            'success': False,
            'error': 'Failed to delete canonical group',
            'message': str(e)
        }), 500


@app.route('/api/canonical-groups/match', methods=['POST'])
def api_match_canonical_entity():
    """
    API endpoint to match an entity name against canonical customer groups

    Request body:
        {
            "entity_name": "Company Name to Match",
            "active_only": true,
            "min_score": 0.75
        }

    Returns: JSON with list of matches sorted by score
    """
    try:
        from canonical_groups_manager import canonical_groups_manager

        data = request.get_json()

        entity_name = data.get('entity_name')
        active_only = data.get('active_only', True)
        min_score = data.get('min_score', 0.75)

        if not entity_name:
            return jsonify({
                'success': False,
                'error': 'Missing required field: entity_name'
            }), 400

        # Match entity
        matches = canonical_groups_manager.match_entity(
            entity_name=entity_name,
            active_only=active_only,
            min_score=min_score
        )

        return jsonify({
            'success': True,
            'entity_name': entity_name,
            'matches': matches,
            'total_matches': len(matches),
            'best_match': matches[0] if matches else None
        })

    except Exception as e:
        app.logger.error(f"Error matching entity: {str(e)}")
        return jsonify({
            'success': False,
            'error': 'Failed to match entity',
            'message': str(e)
        }), 500


################################################################################
# ⛔ LOCKED CODE - DO NOT MODIFY ⛔
# Umbrella Stats API - READ-ONLY FOREVER
# This endpoint ONLY reads from contractor-pay-umbrellas-development table
# NO PayRecords scanning - umbrella company data only
# See: DO-NOT-TOUCH-UMBRELLA-DASHBOARD.md
################################################################################
@app.route('/api/umbrella-stats', methods=['GET'])
def api_umbrella_stats():
    """
    API endpoint to get umbrella company statistics

    Returns: JSON with umbrella companies (NO payment data)
    """
    try:
        from collections import defaultdict
        from datetime import datetime

        app.logger.info("=" * 80)
        app.logger.info("🔍 UMBRELLA STATS API - DETAILED LOGGING")
        app.logger.info("=" * 80)

        # STEP 1: Scan umbrellas table
        app.logger.info("STEP 1: Scanning umbrellas table")
        app.logger.info(f"   Table name: {UMBRELLAS_TABLE}")
        app.logger.info(f"   Table object: {umbrellas_table}")

        umbrellas_response = umbrellas_table.scan(
            FilterExpression='EntityType = :et',
            ExpressionAttributeValues={':et': 'Umbrella'}
        )

        app.logger.info(f"✅ Scan complete")
        app.logger.info(f"   Items returned: {len(umbrellas_response.get('Items', []))}")
        app.logger.info(f"   Scanned count: {umbrellas_response.get('ScannedCount', 0)}")

        app.logger.info("=" * 80)
        app.logger.info("STEP 2: Processing umbrella items into dictionary")
        app.logger.info("=" * 80)

        umbrellas = {}
        for idx, item in enumerate(umbrellas_response.get('Items', [])):
            app.logger.info(f"Processing item {idx + 1}/{len(umbrellas_response.get('Items', []))}")
            app.logger.info(f"   Raw item keys: {list(item.keys())}")

            umbrella_code = item.get('UmbrellaCode')
            app.logger.info(f"   UmbrellaCode: {umbrella_code}")

            if umbrella_code:
                umbrellas[umbrella_code] = {
                    'pk': item.get('PK', ''),
                    'sk': item.get('SK', ''),
                    'umbrella_id': item.get('UmbrellaID', ''),
                    'umbrella_code': item.get('UmbrellaCode', ''),
                    'file_name_variation': item.get('FileNameVariation', ''),
                    'legal_name': item.get('LegalName', ''),
                    'company_reg_no': item.get('CompanyRegNo', ''),
                    'vat_no': item.get('VATNo', ''),
                    'registered_address': item.get('RegisteredAddress', ''),
                    'website': item.get('Website', ''),
                    'primary_contact_email': item.get('PrimaryContactEmail', ''),
                    'contacts': item.get('Contacts', []),
                    'contact_count': item.get('ContactCount', 0),
                    'is_active': item.get('IsActive', True),
                    'entity_type': item.get('EntityType', ''),
                    'metadata': item.get('Metadata', ''),
                    'created_at': item.get('CreatedAt', ''),
                    'updated_at': item.get('UpdatedAt', ''),
                    'gsi1_pk': item.get('GSI1PK', ''),
                    'gsi1_sk': item.get('GSI1SK', ''),
                    'gsi2_pk': item.get('GSI2PK', ''),
                    'gsi2_sk': item.get('GSI2SK', '')
                }
                app.logger.info(f"   ✅ Added {umbrella_code} to umbrellas dictionary")
                app.logger.info(f"      Legal Name: {umbrellas[umbrella_code]['legal_name']}")
                app.logger.info(f"      Contacts: {len(umbrellas[umbrella_code]['contacts'])}")
            else:
                app.logger.warning(f"   ⚠️ Item {idx + 1} has no UmbrellaCode - SKIPPED")

        app.logger.info("=" * 80)
        app.logger.info(f"✅ STEP 2 COMPLETE: Built umbrellas dictionary")
        app.logger.info(f"   Total umbrellas in dictionary: {len(umbrellas)}")
        app.logger.info(f"   Umbrella codes: {list(umbrellas.keys())}")
        app.logger.info("=" * 80)

        # STEP 3: Build result array directly from umbrellas (NO PayRecords scanning)
        app.logger.info("STEP 3: Building result array from umbrella data")
        app.logger.info("   (NO PayRecords scan - umbrella page shows company info only)")

        result = []
        for umbrella_code, data in umbrellas.items():
            umbrella_item = {
                'pk': data['pk'],
                'sk': data['sk'],
                'umbrella_id': data['umbrella_id'],
                'umbrella_code': data['umbrella_code'],
                'file_name_variation': data['file_name_variation'],
                'legal_name': data['legal_name'],
                'company_reg_no': data['company_reg_no'],
                'vat_no': data['vat_no'],
                'registered_address': data['registered_address'],
                'website': data['website'],
                'primary_contact_email': data['primary_contact_email'],
                'contacts': data['contacts'],
                'contact_count': data['contact_count'],
                'is_active': data['is_active'],
                'entity_type': data['entity_type'],
                'metadata': data['metadata'],
                'created_at': data['created_at'],
                'updated_at': data['updated_at'],
                'gsi1_pk': data['gsi1_pk'],
                'gsi1_sk': data['gsi1_sk'],
                'gsi2_pk': data['gsi2_pk'],
                'gsi2_sk': data['gsi2_sk']
            }
            result.append(umbrella_item)

            app.logger.info(f"   ✅ Added {umbrella_code}")
            app.logger.info(f"      Legal: {data['legal_name']}")
            app.logger.info(f"      Reg: {data['company_reg_no']}")
            app.logger.info(f"      VAT: {data['vat_no']}")
            app.logger.info(f"      Contacts: {data['contact_count']}")

        # Sort by umbrella code
        result.sort(key=lambda x: x['umbrella_code'])

        app.logger.info("=" * 80)
        app.logger.info(f"✅ STEP 3 COMPLETE: Built result with {len(result)} umbrellas")
        app.logger.info("=" * 80)

        # Build response
        response_data = {
            'umbrellas': result,
            'summary': {
                'total_umbrellas': len(result)
            }
        }

        app.logger.info("=" * 80)
        app.logger.info("✅ API RESPONSE READY - Returning to client")
        app.logger.info(f"   Response contains {len(result)} umbrella companies")
        app.logger.info("=" * 80)

        return jsonify(response_data)

    except Exception as e:
        app.logger.error(f"Error fetching umbrella stats: {str(e)}")
        import traceback
        app.logger.error(traceback.format_exc())
        return jsonify({'error': 'Failed to fetch umbrella statistics', 'message': str(e)}), 500

################################################################################
# ⛔ END LOCKED CODE - DO NOT MODIFY ABOVE ⛔
################################################################################


@app.route('/api/files/<file_id>/download', methods=['GET'])
def api_download_file(file_id):
    """
    API endpoint to download Excel file from S3

    Returns: Excel file as binary stream or error JSON
    """
    try:
        # Get file metadata from DynamoDB
        response = table.get_item(
            Key={
                'PK': f'FILE#{file_id}',
                'SK': 'METADATA'
            }
        )

        if 'Item' not in response:
            return jsonify({'error': 'File not found'}), 404

        file_item = response['Item']
        s3_key = file_item.get('S3Key')
        original_filename = file_item.get('OriginalFilename', 'file.xlsx')

        if not s3_key:
            return jsonify({'error': 'S3 key not found for file'}), 404

        # Download file from S3
        try:
            s3_response = s3_client.get_object(Bucket=S3_BUCKET, Key=s3_key)
            file_content = s3_response['Body'].read()

            # Return file as binary response
            from flask import make_response
            response = make_response(file_content)
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            response.headers['Content-Disposition'] = f'attachment; filename="{original_filename}"'
            return response

        except ClientError as e:
            app.logger.error(f"S3 error downloading file: {str(e)}")
            return jsonify({'error': 'Failed to download file from S3', 'message': str(e)}), 500

    except Exception as e:
        app.logger.error(f"Error downloading file {file_id}: {str(e)}")
        return jsonify({'error': 'Failed to download file', 'message': str(e)}), 500


@app.route('/api/debug/files', methods=['GET'])
def api_debug_files():
    """
    API endpoint to get list of files for debugging interface

    Returns: JSON with files metadata including navigation info
    """
    try:
        # Query DynamoDB for all File records
        response = table.scan(
            FilterExpression='EntityType = :et',
            ExpressionAttributeValues={':et': 'File'}
        )

        files_list = []
        for item in response.get('Items', []):
            # Convert Decimal to int/float for JSON serialization
            from decimal import Decimal

            def convert_decimal(value):
                if isinstance(value, Decimal):
                    return int(value) if value % 1 == 0 else float(value)
                return value

            file_data = {
                'file_id': item.get('FileID'),
                'filename': item.get('OriginalFilename', 'Unknown'),
                'umbrella': item.get('UmbrellaCode', 'Unknown'),
                'status': item.get('Status', 'UNKNOWN'),
                'uploaded_at': item.get('UploadedAt', item.get('CreatedAt', '')),
                's3_key': item.get('S3Key', ''),
                'file_size': convert_decimal(item.get('FileSizeBytes', 0)),
                'total_records': convert_decimal(item.get('TotalRecords', 0)),
                'valid_records': convert_decimal(item.get('ValidRecords', 0)),
                'error_records': convert_decimal(item.get('ErrorRecords', 0)),
                'error_count': convert_decimal(item.get('ErrorCount', 0)),
                'warning_count': convert_decimal(item.get('WarningCount', 0)),
                'period_id': item.get('PeriodID', 'Unknown'),
                'error_message': item.get('ErrorMessage', '')
            }
            files_list.append(file_data)

        # Sort by upload date (newest first)
        files_list.sort(key=lambda x: x.get('uploaded_at', ''), reverse=True)

        return jsonify({'files': files_list})

    except Exception as e:
        app.logger.error(f"Error fetching debug files: {str(e)}")
        return jsonify({'error': 'Failed to fetch files', 'message': str(e)}), 500


@app.route('/api/files/<file_id>/records', methods=['GET'])
def api_file_records(file_id):
    """
    API endpoint to get PayRecords for a specific file

    Returns: JSON with PayRecords from DynamoDB
    """
    try:
        # Query PayRecords for this file
        response = table.query(
            KeyConditionExpression=Key('PK').eq(f'FILE#{file_id}') & Key('SK').begins_with('RECORD#')
        )

        records = []
        for item in response.get('Items', []):
            # Convert Decimal to float for JSON serialization
            record_data = {}
            for key, value in item.items():
                if isinstance(value, Decimal):
                    record_data[key] = float(value)
                else:
                    record_data[key] = value
            records.append(record_data)

        return jsonify({'records': records, 'count': len(records)})

    except Exception as e:
        app.logger.error(f"Error fetching records for file {file_id}: {str(e)}")
        return jsonify({'error': 'Failed to fetch records', 'message': str(e)}), 500


@app.errorhandler(404)
def not_found(error):
    """404 error handler"""
    return render_template('404.html'), 404


@app.errorhandler(500)
def internal_error(error):
    """500 error handler"""
    return render_template('500.html'), 500


@app.context_processor
def inject_globals():
    """Inject global variables into templates"""
    return {
        'app_name': 'Contractor Pay Tracker',
        'current_year': datetime.now().year
    }


if __name__ == '__main__':
    # Use environment variable for port, or find available port
    # This ensures consistent port across Flask reloader restarts
    if 'FLASK_RUN_PORT' not in os.environ:
        port = find_available_port(start_port=5555)
        os.environ['FLASK_RUN_PORT'] = str(port)
    else:
        port = int(os.environ['FLASK_RUN_PORT'])

    # Only print banner and open browser once (not on reloader restart)
    if os.environ.get('WERKZEUG_RUN_MAIN') != 'true':
        print("\n" + "="*60)
        print("🚀 Contractor Pay Tracker - Flask App")
        print("="*60)
        print(f"\n✅ Server starting on http://127.0.0.1:{port}")
        print(f"✅ Local access: http://localhost:{port}")
        print(f"\n📁 Upload files: http://localhost:{port}/upload")
        print(f"📊 View files: http://localhost:{port}/files")
        print(f"\n⚠️  Press CTRL+C to stop the server")
        print("="*60 + "\n")

        # Open browser after server starts
        def open_browser():
            import time
            time.sleep(2)  # Wait for server to start
            webbrowser.open(f'http://localhost:{port}')

        threading.Thread(target=open_browser, daemon=True).start()

    app.run(
        host='127.0.0.1',
        port=port,
        debug=True,
        use_reloader=True
    )
