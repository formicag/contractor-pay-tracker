"""
Excel File Processing Utilities

Functions for extracting contractor data from pay files
"""

import openpyxl
from io import BytesIO
from typing import List, Set

def extract_contractor_names(file_content: bytes) -> List[str]:
    """
    Extract unique contractor names from Excel file

    Expected format:
    - Row 1: Headers (employee id, surname, forename, ...)
    - Row 2+: Data rows with surname in column B and forename in column C

    Args:
        file_content: Excel file as bytes

    Returns:
        List of unique contractor names in "Forename Surname" format, sorted alphabetically
    """
    try:
        # Load workbook from bytes
        wb = openpyxl.load_workbook(BytesIO(file_content), read_only=True, data_only=False)

        # Get first sheet
        ws = wb[wb.sheetnames[0]]

        # Extract names from rows 2+ (skip header row)
        names_set: Set[str] = set()

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 3:
                continue

            # Column B (index 1) = surname, Column C (index 2) = forename
            surname = row[1]
            forename = row[2]

            # Skip if either is missing
            if not surname or not forename:
                continue

            # Clean and format name
            surname_clean = str(surname).strip()
            forename_clean = str(forename).strip()

            if surname_clean and forename_clean:
                full_name = f"{forename_clean} {surname_clean}"
                names_set.add(full_name)

        wb.close()

        # Return sorted list
        return sorted(list(names_set))

    except Exception as e:
        print(f"Error extracting contractor names: {str(e)}")
        return []


def get_contractor_summary(names: List[str]) -> str:
    """
    Create a summary string of contractor names

    Args:
        names: List of contractor names

    Returns:
        Summary string (e.g., "3 contractors: Alice Smith, Bob Jones, Charlie Brown")
    """
    if not names:
        return "Unknown"

    count = len(names)
    if count == 1:
        return names[0]
    elif count <= 3:
        return f"{count} contractors: {', '.join(names)}"
    else:
        return f"{count} contractors: {', '.join(names[:3])}, ..."


def validate_payfile_structure(file_content: bytes) -> dict:
    """
    Validate Excel file has expected structure for pay files

    Args:
        file_content: Excel file as bytes

    Returns:
        Dict with validation results: {'valid': bool, 'message': str, 'row_count': int}
    """
    try:
        wb = openpyxl.load_workbook(BytesIO(file_content), read_only=True, data_only=False)
        ws = wb[wb.sheetnames[0]]

        # Check if file has data
        if ws.max_row < 2:
            return {
                'valid': False,
                'message': 'File appears to be empty (no data rows)',
                'row_count': 0
            }

        # Check header row has expected columns
        header_row = [cell for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]

        expected_columns = ['employee id', 'surname', 'forename']
        has_expected = all(
            any(str(col).lower().strip() == exp.lower() for col in header_row if col)
            for exp in expected_columns
        )

        if not has_expected:
            return {
                'valid': False,
                'message': f'Missing expected columns. Found: {header_row[:5]}',
                'row_count': ws.max_row - 1
            }

        row_count = ws.max_row - 1  # Exclude header

        wb.close()

        return {
            'valid': True,
            'message': f'Valid payfile with {row_count} data rows',
            'row_count': row_count
        }

    except Exception as e:
        return {
            'valid': False,
            'message': f'Error validating file: {str(e)}',
            'row_count': 0
        }
