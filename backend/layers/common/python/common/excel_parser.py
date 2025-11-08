"""
Excel file parser for contractor pay files
Handles the specific format from umbrella companies
"""

print("[EXCEL_PARSER_MODULE] Starting excel_parser.py module load")

import re
from datetime import datetime
from decimal import Decimal
from typing import Dict, List, Optional

print("[EXCEL_PARSER_MODULE] Imported standard library modules: re, datetime, Decimal, typing")

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

print("[EXCEL_PARSER_MODULE] Imported openpyxl modules")


class PayFileParser:
    """Parse contractor pay Excel files"""

    def __init__(self, file_path: str):
        print(f"[PAYFILEPARSER_INIT] Starting PayFileParser initialization with file_path={file_path}")

        self.file_path = file_path
        print(f"[PAYFILEPARSER_INIT] Set self.file_path={self.file_path}")

        print(f"[PAYFILEPARSER_INIT] Loading workbook from {file_path}")
        self.workbook = openpyxl.load_workbook(file_path, data_only=True)
        print(f"[PAYFILEPARSER_INIT] Workbook loaded: {self.workbook}")

        self.worksheet = self.workbook.active
        print(f"[PAYFILEPARSER_INIT] Active worksheet: {self.worksheet}")
        print(f"[PAYFILEPARSER_INIT] PayFileParser initialization complete")

    def extract_metadata(self) -> Dict:
        """Extract metadata from filename and file content"""
        print(f"[EXTRACT_METADATA] Starting metadata extraction")

        filename = self.file_path.split('/')[-1]
        print(f"[EXTRACT_METADATA] Extracted filename from path: {filename}")

        umbrella_code = self._extract_umbrella_code()
        submission_date = self._extract_submission_date()

        metadata = {
            'filename': filename,
            'umbrella_code': umbrella_code,
            'submission_date': submission_date
        }
        print(f"[EXTRACT_METADATA] Final metadata: {metadata}")
        return metadata

    def find_header_row(self) -> int:
        """Find the row containing column headers"""
        print(f"[FIND_HEADER_ROW] Starting header row search")

        # Common header variations
        header_indicators = ['employee id', 'surname', 'forename', 'unit', 'rate', 'amount']
        print(f"[FIND_HEADER_ROW] Header indicators: {header_indicators}")

        print(f"[FIND_HEADER_ROW] Iterating through rows 1-20")
        for row_idx, row in enumerate(self.worksheet.iter_rows(min_row=1, max_row=20), start=1):
            print(f"[FIND_HEADER_ROW] Checking row {row_idx}")

            row_text = ' '.join([str(cell.value or '').lower() for cell in row])
            print(f"[FIND_HEADER_ROW] Row {row_idx} text: {row_text[:100]}...")

            # Check if this row contains most header indicators
            matches = sum(1 for indicator in header_indicators if indicator in row_text)
            print(f"[FIND_HEADER_ROW] Row {row_idx} has {matches} header indicator matches")

            if matches >= 4:
                print(f"[FIND_HEADER_ROW] Found header row at index {row_idx} with {matches} matches")
                return row_idx

        print(f"[FIND_HEADER_ROW] No header row found, defaulting to row 1")
        return 1  # Default to first row

    def parse_records(self) -> List[Dict]:
        """Parse all pay records from the Excel file"""
        print(f"[PARSE_RECORDS] Starting record parsing")

        header_row = self.find_header_row()
        print(f"[PARSE_RECORDS] Header row identified: {header_row}")

        # Get column mapping
        print(f"[PARSE_RECORDS] Getting column mapping")
        column_map = self._get_column_mapping(header_row)
        print(f"[PARSE_RECORDS] Column mapping: {column_map}")

        records = []
        print(f"[PARSE_RECORDS] Initialized empty records list")

        print(f"[PARSE_RECORDS] Starting row iteration from row {header_row + 1}")

        for row_idx, row in enumerate(self.worksheet.iter_rows(min_row=header_row + 1), start=header_row + 1):
            row_number = row_idx
            print(f"[PARSE_RECORDS] Processing row {row_idx}, row_number={row_number}")

            # Skip empty rows
            is_empty = all(cell.value is None or str(cell.value).strip() == '' for cell in row)
            print(f"[PARSE_RECORDS] Row {row_idx} is_empty={is_empty}")
            if is_empty:
                print(f"[PARSE_RECORDS] Skipping empty row {row_idx}")
                continue

            # Check if this is a duplicate header row
            first_cell = str(row[0].value or '').lower()
            print(f"[PARSE_RECORDS] Row {row_idx} first cell: {first_cell}")
            if 'employee' in first_cell or 'surname' in first_cell:
                print(f"[PARSE_RECORDS] Row {row_idx} is duplicate header, skipping")
                continue

            try:
                print(f"[PARSE_RECORDS] Attempting to parse row {row_idx}")
                record = self._parse_row(row, column_map, row_number, row_idx)
                print(f"[PARSE_RECORDS] Parsed record: {record}")

                if record:
                    print(f"[PARSE_RECORDS] Record is valid, appending to records list")
                    records.append(record)
                    print(f"[PARSE_RECORDS] Total records so far: {len(records)}")
                else:
                    print(f"[PARSE_RECORDS] Record is None, skipping")
            except Exception as e:
                # Log parsing error but continue
                print(f"[PARSE_RECORDS] ERROR parsing row {row_idx}: {type(e).__name__}: {str(e)}")
                print(f"[PARSE_RECORDS] Continuing to next row")
                continue

        print(f"[PARSE_RECORDS] Parsing complete. Total records parsed: {len(records)}")
        return records

    def _get_column_mapping(self, header_row: int) -> Dict[str, int]:
        """Map column names to their indices"""
        print(f"[GET_COLUMN_MAPPING] Starting column mapping for header_row={header_row}")

        headers = []
        print(f"[GET_COLUMN_MAPPING] Reading headers from row {header_row}")
        for idx, cell in enumerate(self.worksheet[header_row]):
            header_text = str(cell.value or '').lower().strip()
            print(f"[GET_COLUMN_MAPPING] Column {idx}: '{header_text}'")
            headers.append(header_text)

        print(f"[GET_COLUMN_MAPPING] All headers: {headers}")

        # Map common variations to standard names
        column_map = {}
        print(f"[GET_COLUMN_MAPPING] Mapping headers to standard names")

        for idx, header in enumerate(headers):
            print(f"[GET_COLUMN_MAPPING] Processing header {idx}: '{header}'")

            if 'employee' in header and 'id' in header:
                print(f"[GET_COLUMN_MAPPING] Matched 'employee_id' at index {idx}")
                column_map['employee_id'] = idx
            elif 'surname' in header or 'last' in header:
                print(f"[GET_COLUMN_MAPPING] Matched 'surname' at index {idx}")
                column_map['surname'] = idx
            elif 'forename' in header or 'first' in header:
                print(f"[GET_COLUMN_MAPPING] Matched 'forename' at index {idx}")
                column_map['forename'] = idx
            elif header == 'unit' or 'days' in header:
                print(f"[GET_COLUMN_MAPPING] Matched 'unit' at index {idx}")
                column_map['unit'] = idx
            elif 'rate' in header.lower() and 'day' in header.lower():
                print(f"[GET_COLUMN_MAPPING] Matched 'rate' at index {idx}")
                column_map['rate'] = idx
            elif header == 'per':
                print(f"[GET_COLUMN_MAPPING] Matched 'per' at index {idx}")
                column_map['per'] = idx
            elif header == 'amount' and 'vat' not in header:
                print(f"[GET_COLUMN_MAPPING] Matched 'amount' at index {idx}")
                column_map['amount'] = idx
            elif 'vat' in header and 'amount' not in header:
                print(f"[GET_COLUMN_MAPPING] Matched 'vat' at index {idx}")
                column_map['vat'] = idx
            elif 'total' in header and 'hours' in header:
                print(f"[GET_COLUMN_MAPPING] Matched 'total_hours' at index {idx}")
                column_map['total_hours'] = idx
            elif 'company' in header:
                print(f"[GET_COLUMN_MAPPING] Matched 'company' at index {idx}")
                column_map['company'] = idx
            elif 'notes' in header or 'description' in header:
                print(f"[GET_COLUMN_MAPPING] Matched 'notes' at index {idx}")
                column_map['notes'] = idx
            else:
                print(f"[GET_COLUMN_MAPPING] No match for header '{header}' at index {idx}")

        print(f"[GET_COLUMN_MAPPING] Final column map: {column_map}")
        return column_map

    def _parse_row(self, row, column_map: Dict[str, int], row_number: int, row_idx: int) -> Optional[Dict]:
        """Parse a single row into a pay record"""
        print(f"[PARSE_ROW] Starting row parse for row_number={row_number}, row_idx={row_idx}")

        def get_cell_value(col_name: str, default=None):
            """Safely get cell value"""
            print(f"[PARSE_ROW_GET_CELL] Getting value for col_name='{col_name}', default={default}")

            if col_name not in column_map:
                print(f"[PARSE_ROW_GET_CELL] Column '{col_name}' not in column_map, returning default")
                return default

            idx = column_map[col_name]
            print(f"[PARSE_ROW_GET_CELL] Column '{col_name}' is at index {idx}")

            if idx >= len(row):
                print(f"[PARSE_ROW_GET_CELL] Index {idx} >= row length {len(row)}, returning default")
                return default

            value = row[idx].value
            print(f"[PARSE_ROW_GET_CELL] Cell value: {value}")
            result = value if value is not None else default
            print(f"[PARSE_ROW_GET_CELL] Returning: {result}")
            return result

        # Get basic fields
        print(f"[PARSE_ROW] Getting employee_id")
        employee_id = str(get_cell_value('employee_id', '')).strip()
        print(f"[PARSE_ROW] employee_id={employee_id}")

        print(f"[PARSE_ROW] Getting surname")
        surname = str(get_cell_value('surname', '')).strip()
        print(f"[PARSE_ROW] surname={surname}")

        print(f"[PARSE_ROW] Getting forename")
        forename = str(get_cell_value('forename', '')).strip()
        print(f"[PARSE_ROW] forename={forename}")

        # Skip rows without essential data
        if not surname or not forename:
            print(f"[PARSE_ROW] Missing surname or forename, returning None")
            return None

        print(f"[PARSE_ROW] Essential data present, continuing parse")

        # Get numeric fields
        print(f"[PARSE_ROW] Getting unit_days")
        unit_days = self._parse_decimal(get_cell_value('unit', 0))
        print(f"[PARSE_ROW] unit_days={unit_days}")

        print(f"[PARSE_ROW] Getting day_rate")
        day_rate = self._parse_decimal(get_cell_value('rate', 0))
        print(f"[PARSE_ROW] day_rate={day_rate}")

        print(f"[PARSE_ROW] Getting amount")
        amount = self._parse_decimal(get_cell_value('amount', 0))
        print(f"[PARSE_ROW] amount={amount}")

        print(f"[PARSE_ROW] Getting vat_amount")
        vat_amount = self._parse_decimal(get_cell_value('vat', 0))
        print(f"[PARSE_ROW] vat_amount={vat_amount}")

        print(f"[PARSE_ROW] Getting total_hours")
        total_hours = self._parse_decimal(get_cell_value('total_hours', 0))
        print(f"[PARSE_ROW] total_hours={total_hours}")

        # Get other fields
        print(f"[PARSE_ROW] Getting notes")
        notes = str(get_cell_value('notes', '')).strip()
        print(f"[PARSE_ROW] notes={notes}")

        print(f"[PARSE_ROW] Getting company")
        company = str(get_cell_value('company', '')).strip()
        print(f"[PARSE_ROW] company={company}")

        # Calculate gross amount (amount + vat)
        print(f"[PARSE_ROW] Calculating gross_amount = amount ({amount}) + vat_amount ({vat_amount})")
        gross_amount = amount + vat_amount
        print(f"[PARSE_ROW] gross_amount={gross_amount}")

        # Determine record type
        print(f"[PARSE_ROW] Determining record type from notes")
        record_type = 'NORMAL'
        if self._is_overtime_record(notes):
            print(f"[PARSE_ROW] Found 'overtime' in notes, setting record_type=OVERTIME")
            record_type = 'OVERTIME'
        elif 'expense' in notes.lower():
            print(f"[PARSE_ROW] Found 'expense' in notes, setting record_type=EXPENSE")
            record_type = 'EXPENSE'
        else:
            print(f"[PARSE_ROW] No special keywords in notes, record_type=STANDARD")

        # Calculate final total_hours if not provided
        print(f"[PARSE_ROW] Calculating final total_hours")
        final_total_hours = float(total_hours) if total_hours else float(unit_days * Decimal('8'))
        print(f"[PARSE_ROW] final_total_hours={final_total_hours}")

        record = {
            'row_number': row_number,
            'row_idx': row_idx,
            'employee_id': employee_id,
            'surname': surname,
            'forename': forename,
            'unit_days': float(unit_days),
            'day_rate': float(day_rate),
            'amount': float(amount),
            'vat_amount': float(vat_amount),
            'gross_amount': float(gross_amount),
            'total_hours': final_total_hours,
            'record_type': record_type,
            'notes': notes,
            'company': company
        }
        print(f"[PARSE_ROW] Final record: {record}")
        return record

    @staticmethod
    def _parse_decimal(value) -> Decimal:
        """Parse value to Decimal, handling various formats"""
        print(f"[PARSE_DECIMAL] Parsing value: {value} (type: {type(value).__name__})")

        if value is None:
            print(f"[PARSE_DECIMAL] Value is None, returning Decimal('0')")
            return Decimal('0')

        if isinstance(value, (int, float)):
            print(f"[PARSE_DECIMAL] Value is int or float, converting to Decimal")
            result = Decimal(str(value))
            print(f"[PARSE_DECIMAL] Converted to: {result}")
            return result

        # Handle string values
        print(f"[PARSE_DECIMAL] Value is not int/float, treating as string")
        str_value = str(value).strip().replace(',', '').replace('£', '')
        print(f"[PARSE_DECIMAL] Cleaned string value: '{str_value}'")

        if not str_value or str_value == '-':
            print(f"[PARSE_DECIMAL] String value is empty or '-', returning Decimal('0')")
            return Decimal('0')

        try:
            print(f"[PARSE_DECIMAL] Attempting to convert '{str_value}' to Decimal")
            result = Decimal(str_value)
            print(f"[PARSE_DECIMAL] Successfully converted to: {result}")
            return result
        except Exception as e:
            print(f"[PARSE_DECIMAL] Conversion failed: {type(e).__name__}: {str(e)}")
            print(f"[PARSE_DECIMAL] Returning Decimal('0')")
            return Decimal('0')

    def _is_overtime_record(self, notes: str) -> bool:
        """Check if a record is an overtime record based on notes"""
        print(f"[IS_OVERTIME_RECORD] Checking if notes contain 'overtime' or 'OT': {notes}")
        result = 'overtime' in notes.lower() or notes.upper() == 'OT'
        print(f"[IS_OVERTIME_RECORD] Result: {result}")
        return result

    def _extract_umbrella_code(self) -> Optional[str]:
        """Extract umbrella company code from filename"""
        filename = self.file_path.split('/')[-1]
        print(f"[EXTRACT_UMBRELLA_CODE] Extracted filename: {filename}")

        umbrella_patterns = {
            'NASA': r'NASA',
            'PAYSTREAM': r'PAYSTREAM',
            'PARASOL': r'Parasol',
            'CLARITY': r'Clarity',
            'GIANT': r'GIANT',
            'WORKWELL': r'WORKWELL'
        }
        print(f"[EXTRACT_UMBRELLA_CODE] Umbrella patterns: {umbrella_patterns}")

        umbrella_code = None
        print(f"[EXTRACT_UMBRELLA_CODE] Searching for umbrella company in filename")
        for code, pattern in umbrella_patterns.items():
            print(f"[EXTRACT_UMBRELLA_CODE] Checking pattern '{pattern}' for code '{code}'")
            if re.search(pattern, filename, re.IGNORECASE):
                print(f"[EXTRACT_UMBRELLA_CODE] Match found! Setting umbrella_code={code}")
                umbrella_code = code
                break

        if umbrella_code is None:
            print(f"[EXTRACT_UMBRELLA_CODE] No umbrella company match found")
        else:
            print(f"[EXTRACT_UMBRELLA_CODE] Final umbrella_code={umbrella_code}")

        return umbrella_code

    def _extract_submission_date(self) -> Optional[str]:
        """Extract submission date from filename (DDMMYYYY format)"""
        filename = self.file_path.split('/')[-1]
        print(f"[EXTRACT_SUBMISSION_DATE] Extracted filename: {filename}")

        print(f"[EXTRACT_SUBMISSION_DATE] Searching for date in filename (DDMMYYYY format)")
        date_match = re.search(r'(\d{8})', filename)
        print(f"[EXTRACT_SUBMISSION_DATE] Date regex match result: {date_match}")

        submission_date = None
        if date_match:
            date_str = date_match.group(1)
            print(f"[EXTRACT_SUBMISSION_DATE] Extracted date string: {date_str}")
            submission_date = date_str
            print(f"[EXTRACT_SUBMISSION_DATE] Parsed submission_date: {submission_date}")
        else:
            print(f"[EXTRACT_SUBMISSION_DATE] No date found in filename")

        return submission_date

    def __enter__(self):
        """Context manager entry"""
        print(f"[CONTEXT_MANAGER] Entering context manager")
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        """Context manager exit"""
        print(f"[CONTEXT_MANAGER] Exiting context manager, exc_type={exc_type}")
        self.close()
        return False

    def close(self):
        """Close the workbook"""
        print(f"[CLOSE] Closing workbook")
        self.workbook.close()
        print(f"[CLOSE] Workbook closed")

print("[EXCEL_PARSER_MODULE] excel_parser.py module load complete")
